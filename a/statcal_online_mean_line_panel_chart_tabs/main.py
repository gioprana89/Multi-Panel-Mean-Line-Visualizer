import math
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="STATCAL ONLINE - Panel Mean Line Chart Builder",
    page_icon="📈",
    layout="wide",
)


# =====================================================
# CONSTANTS
# =====================================================

APP_NAME = "STATCAL ONLINE"
APP_TITLE = "Panel Mean Line Chart Builder"
APP_UPDATED = "Last updated on June 14, 2026"
WEBSITE_URL = "https://statcal.com/"
TRAINING_DATA_URL = "https://drive.google.com/drive/folders/1lg3RDAAfX7ua-zwCTktSqkd4Oa36yZ0R?usp=sharing"
LOGO_PATH = Path("logo_statcal.png")
SAMPLE_DATA_PATH = Path("data_sektor_energi.xlsx")
BRAND_COLOR = "1F4E79"

STATISTICS_COLUMNS = ["N", "Minimum", "Maximum", "Mean", "Standard Deviation"]

COLOR_PALETTES = {
    "Scopus Blue Orange": ["#1F4E79", "#E97132", "#70AD47", "#FFC000", "#7030A0", "#00A6A6", "#C00000", "#595959"],
    "Nature Publication": ["#1B7837", "#762A83", "#5AAE61", "#9970AB", "#A6DBA0", "#C2A5CF", "#00441B", "#40004B"],
    "Finance Professional": ["#0B3C49", "#3282B8", "#BBE1FA", "#F9A03F", "#D1495B", "#00798C", "#30638E", "#003D5B"],
    "Economics Journal": ["#003F5C", "#BC5090", "#FFA600", "#58508D", "#FF6361", "#2F4B7C", "#A05195", "#D45087"],
    "Black Gray Academic": ["#000000", "#404040", "#666666", "#808080", "#A6A6A6", "#BFBFBF", "#595959", "#262626"],
    "High Contrast Q1": ["#2166AC", "#B2182B", "#4D9221", "#762A83", "#D6604D", "#1B7837", "#4393C3", "#9970AB"],
}

THEMES = {
    "White Publication": {
        "figure_facecolor": "white",
        "axes_facecolor": "white",
        "text_color": "#111111",
        "grid_color": "#D9D9D9",
        "spine_color": "#222222",
    },
    "Light Gray Editorial": {
        "figure_facecolor": "#F7F7F7",
        "axes_facecolor": "#FFFFFF",
        "text_color": "#111111",
        "grid_color": "#D0D0D0",
        "spine_color": "#333333",
    },
    "Warm Ivory Journal": {
        "figure_facecolor": "#FBF7EF",
        "axes_facecolor": "#FFFDF8",
        "text_color": "#1F1F1F",
        "grid_color": "#DDD4C4",
        "spine_color": "#3A3A3A",
    },
    "Cool Blue Scientific": {
        "figure_facecolor": "#F3F7FB",
        "axes_facecolor": "#FFFFFF",
        "text_color": "#0B1F33",
        "grid_color": "#C8D6E5",
        "spine_color": "#1F4E79",
    },
    "Dark Navy Presentation": {
        "figure_facecolor": "#0B1320",
        "axes_facecolor": "#111C2E",
        "text_color": "#FFFFFF",
        "grid_color": "#3B4A5F",
        "spine_color": "#B8C7D9",
    },
}

MARKERS = ["o", "s", "^", "D", "P", "X", "v", "<", ">", "*"]


# =====================================================
# GENERAL HELPERS
# =====================================================

def safe_streamlit_image(image_path: Path, width: int = 220) -> None:
    try:
        st.image(str(image_path), width=width)
    except Exception:
        st.markdown("### STATCAL")


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [re.sub(r"\s+", " ", str(col)).strip() for col in df.columns]
    df = df.dropna(how="all")

    unnamed_cols = [col for col in df.columns if str(col).lower().startswith("unnamed")]
    for col in unnamed_cols:
        if df[col].isna().all():
            df = df.drop(columns=[col])
    return df


def make_arrow_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    safe_df = df.copy()
    for col in safe_df.columns:
        if safe_df[col].dtype == "object":
            safe_df[col] = safe_df[col].astype(str).replace({"nan": "", "None": "", "NaT": ""})
    return safe_df


def to_numeric_series(series: pd.Series) -> pd.Series:
    """
    Convert a pandas Series to numeric.
    Supports comma separators, unicode minus signs, parentheses for negatives,
    percentages, and suffixes such as K, M, B, and T.
    """
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")

    text = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("−", "-", regex=False)
        .str.replace("—", "", regex=False)
        .str.strip()
    )

    text = text.replace({
        "": pd.NA,
        "nan": pd.NA,
        "None": pd.NA,
        "NaT": pd.NA,
        "-": pd.NA,
        "N/A": pd.NA,
        "NA": pd.NA,
        "n/a": pd.NA,
        "na": pd.NA,
    })

    def convert_value(value):
        if pd.isna(value):
            return np.nan

        value = str(value).strip()
        if not value:
            return np.nan

        negative = False
        if value.startswith("(") and value.endswith(")"):
            negative = True
            value = value[1:-1].strip()

        multiplier = 1.0
        last_char = value[-1:].lower()
        if last_char == "k":
            multiplier = 1_000.0
            value = value[:-1]
        elif last_char == "m":
            multiplier = 1_000_000.0
            value = value[:-1]
        elif last_char == "b":
            multiplier = 1_000_000_000.0
            value = value[:-1]
        elif last_char == "t":
            multiplier = 1_000_000_000_000.0
            value = value[:-1]

        value = value.replace("%", "").strip()

        try:
            number = float(value) * multiplier
            return -number if negative else number
        except Exception:
            return np.nan

    return text.map(convert_value)


def detect_numeric_columns(df: pd.DataFrame, min_valid_ratio: float = 0.45) -> List[str]:
    numeric_cols = []
    for col in df.columns:
        s = df[col]
        non_null = s.notna().sum()
        if non_null == 0:
            continue
        numeric_s = to_numeric_series(s)
        valid_ratio = numeric_s.notna().sum() / max(non_null, 1)
        if valid_ratio >= min_valid_ratio:
            numeric_cols.append(col)
    return numeric_cols


def default_numeric_columns(columns: List[str]) -> List[str]:
    excluded = {"year", "tahun", "date", "time", "period", "periode"}
    preferred = [col for col in columns if str(col).strip().lower() not in excluded]
    return preferred[:3] if preferred else columns[:3]


def sorted_unique_values(series: pd.Series) -> List:
    values = series.dropna().unique().tolist()
    try:
        return sorted(values)
    except Exception:
        return sorted(values, key=lambda x: str(x).lower())


def default_filter_columns(columns: List[str]) -> List[str]:
    preferred_terms = ["company", "source code", "ticker", "code", "year", "tahun", "sector", "sektor"]
    defaults = []
    for term in preferred_terms:
        for col in columns:
            lower = str(col).strip().lower()
            if term == lower or term in lower:
                if col not in defaults:
                    defaults.append(col)
    return defaults[:3]


def preferred_option(options: List[str], candidates: List[str], fallback_index: int = 0) -> int:
    for candidate in candidates:
        for idx, option in enumerate(options):
            if str(option).strip().lower() == candidate.lower():
                return idx
    return fallback_index


def apply_category_filters(df: pd.DataFrame, filter_cols: List[str]) -> pd.DataFrame:
    filtered = df.copy()
    for col in filter_cols:
        if col not in filtered.columns:
            continue

        values = sorted_unique_values(filtered[col])
        if not values:
            continue

        selected_values = st.sidebar.multiselect(
            f"Select values for: {col}",
            options=values,
            default=values,
            format_func=lambda x: str(x),
        )

        if selected_values:
            filtered = filtered[filtered[col].isin(selected_values)]
        else:
            filtered = filtered.iloc[0:0]

    return filtered


def prepare_numeric_dataframe(df: pd.DataFrame, numeric_cols: List[str]) -> pd.DataFrame:
    numeric_df = pd.DataFrame(index=df.index)
    for col in numeric_cols:
        numeric_df[col] = to_numeric_series(df[col])
    return numeric_df


def get_palette_color_list(palette_name: str) -> List[str]:
    return COLOR_PALETTES.get(palette_name, COLOR_PALETTES["Scopus Blue Orange"])


def get_theme(theme_name: str) -> Dict[str, str]:
    return THEMES.get(theme_name, THEMES["White Publication"])


def format_x_labels(values: List) -> List[str]:
    labels = []
    for value in values:
        if pd.isna(value):
            labels.append("Missing")
        elif isinstance(value, pd.Timestamp):
            labels.append(value.strftime("%Y-%m-%d"))
        else:
            labels.append(str(value))
    return labels


def sort_values_safely(values: List) -> List:
    try:
        return sorted(values)
    except Exception:
        return sorted(values, key=lambda x: str(x).lower())


def format_numeric_label(value: float, decimal_digits: int, compact: bool = False) -> str:
    if pd.isna(value) or not np.isfinite(value):
        return ""
    if compact:
        abs_value = abs(value)
        if abs_value >= 1_000_000_000_000:
            return f"{value / 1_000_000_000_000:.{decimal_digits}f}T"
        if abs_value >= 1_000_000_000:
            return f"{value / 1_000_000_000:.{decimal_digits}f}B"
        if abs_value >= 1_000_000:
            return f"{value / 1_000_000:.{decimal_digits}f}M"
        if abs_value >= 1_000:
            return f"{value / 1_000:.{decimal_digits}f}K"
    return f"{value:.{decimal_digits}f}"


# =====================================================
# DESCRIPTIVE STATISTICS
# =====================================================

@st.cache_data(ttl=3600, max_entries=20)
def compute_descriptive_statistics(df: pd.DataFrame, numeric_cols: List[str], decimal_digits: int) -> pd.DataFrame:
    numeric_df = prepare_numeric_dataframe(df, numeric_cols)
    rows = []
    for col in numeric_cols:
        s = numeric_df[col].dropna()
        rows.append({
            "Variable": col,
            "N": int(s.count()),
            "Minimum": s.min() if len(s) else np.nan,
            "Maximum": s.max() if len(s) else np.nan,
            "Mean": s.mean() if len(s) else np.nan,
            "Standard Deviation": s.std(ddof=1) if len(s) > 1 else np.nan,
        })
    result = pd.DataFrame(rows)
    for col in STATISTICS_COLUMNS[1:]:
        if col in result.columns:
            result[col] = result[col].round(decimal_digits)
    return result


@st.cache_data(ttl=3600, max_entries=20)
def compute_grouped_descriptive_statistics(
    df: pd.DataFrame,
    numeric_cols: List[str],
    group_cols: List[str],
    decimal_digits: int,
) -> pd.DataFrame:
    if not group_cols:
        return compute_descriptive_statistics(df, numeric_cols, decimal_digits)

    rows = []
    for numeric_col in numeric_cols:
        temp = df[group_cols].copy()
        temp["__value__"] = to_numeric_series(df[numeric_col])
        grouped = temp.groupby(group_cols, dropna=False)["__value__"]
        stats = grouped.agg(
            N="count",
            Minimum="min",
            Maximum="max",
            Mean="mean",
            **{"Standard Deviation": "std"},
        ).reset_index()
        stats.insert(len(group_cols), "Variable", numeric_col)
        rows.append(stats)

    result = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    for col in STATISTICS_COLUMNS[1:]:
        if col in result.columns:
            result[col] = result[col].round(decimal_digits)
    return result


# =====================================================
# LINE CHART DATA
# =====================================================

@st.cache_data(ttl=3600, max_entries=20)
def build_mean_line_long_data(
    df: pd.DataFrame,
    x_col: str,
    numeric_cols: List[str],
    split_col: Optional[str],
    sort_x: bool,
) -> pd.DataFrame:
    rows = []

    for variable in numeric_cols:
        if variable not in df.columns or x_col not in df.columns:
            continue

        cols = [x_col, variable]
        use_split = split_col and split_col != "None" and split_col in df.columns
        if use_split:
            cols.append(split_col)

        temp = df[cols].copy()
        temp[variable] = to_numeric_series(temp[variable])
        temp = temp.dropna(subset=[x_col, variable])

        if temp.empty:
            continue

        if use_split:
            grouped = temp.groupby([x_col, split_col], dropna=False)[variable].mean().reset_index()
            grouped = grouped.rename(columns={variable: "Mean"})
            grouped["Split"] = grouped[split_col].fillna("Missing").astype(str)
        else:
            grouped = temp.groupby(x_col, dropna=False)[variable].mean().reset_index()
            grouped = grouped.rename(columns={variable: "Mean"})
            grouped["Split"] = variable

        grouped["Variable"] = variable
        rows.append(grouped[[x_col, "Variable", "Split", "Mean"]])

    if not rows:
        return pd.DataFrame(columns=[x_col, "Variable", "Split", "Mean"])

    result = pd.concat(rows, ignore_index=True)
    if sort_x:
        try:
            result = result.sort_values(["Variable", "Split", x_col]).reset_index(drop=True)
        except Exception:
            result["__x_sort__"] = result[x_col].astype(str)
            result = result.sort_values(["Variable", "Split", "__x_sort__"]).drop(columns="__x_sort__").reset_index(drop=True)
    return result


def pivot_for_variable(long_df: pd.DataFrame, x_col: str, variable: str) -> pd.DataFrame:
    temp = long_df[long_df["Variable"] == variable].copy()
    if temp.empty:
        return pd.DataFrame()
    pivot = temp.pivot_table(index=x_col, columns="Split", values="Mean", aggfunc="mean")
    try:
        pivot = pivot.sort_index()
    except Exception:
        ordered = sorted(pivot.index.tolist(), key=lambda x: str(x).lower())
        pivot = pivot.loc[ordered]
    return pivot


# =====================================================
# EXCEL EXPORT
# =====================================================

def style_excel_workbook(workbook) -> None:
    header_fill = PatternFill(start_color=BRAND_COLOR, end_color=BRAND_COLOR, fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)


def dataframe_to_excel_bytes(
    descriptive_stats: pd.DataFrame,
    grouped_descriptive_stats: pd.DataFrame,
    chart_long_data: pd.DataFrame,
    filtered_df: pd.DataFrame,
    metadata: Dict,
) -> bytes:
    output = BytesIO()
    metadata_df = pd.DataFrame([{"Item": key, "Value": value} for key, value in metadata.items()])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False)
        descriptive_stats.to_excel(writer, sheet_name="Descriptive Statistics", index=False)
        grouped_descriptive_stats.to_excel(writer, sheet_name="Grouped Descriptive", index=False)
        chart_long_data.to_excel(writer, sheet_name="Mean Line Chart Data", index=False)
        filtered_df.to_excel(writer, sheet_name="Filtered Data", index=False)
        style_excel_workbook(writer.book)

    output.seek(0)
    return output.getvalue()


# =====================================================
# FIGURE EXPORT AND CHART STYLE
# =====================================================

def fig_to_png_bytes(fig, dpi: int, transparent_background: bool = False) -> bytes:
    buffer = BytesIO()
    fig.savefig(
        buffer,
        format="png",
        dpi=dpi,
        bbox_inches="tight",
        facecolor="none" if transparent_background else fig.get_facecolor(),
        transparent=transparent_background,
    )
    buffer.seek(0)
    return buffer.getvalue()


def apply_common_chart_style(ax, theme: Dict[str, str], show_grid: bool) -> None:
    ax.set_facecolor(theme["axes_facecolor"])
    ax.tick_params(axis="both", colors=theme["text_color"])
    ax.xaxis.label.set_color(theme["text_color"])
    ax.yaxis.label.set_color(theme["text_color"])
    ax.title.set_color(theme["text_color"])

    if show_grid:
        ax.grid(axis="both", linestyle="--", linewidth=0.6, alpha=0.45, color=theme["grid_color"])
        ax.set_axisbelow(True)

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_color(theme["spine_color"])
    ax.spines["left"].set_color(theme["spine_color"])


def create_panel_mean_line_chart(
    long_df: pd.DataFrame,
    x_col: str,
    panel_variables: List[str],
    title: str,
    subtitle: str,
    x_label: str,
    y_label: str,
    figure_width: float,
    figure_height: float,
    panel_columns: int,
    share_y_axis: bool,
    theme_name: str,
    font_family: str,
    line_colors: Dict[str, str],
    line_width: float,
    show_markers: bool,
    marker_style: str,
    marker_size: int,
    show_value_labels: bool,
    value_label_decimal_digits: int,
    compact_labels: bool,
    title_font_size: int,
    subtitle_font_size: int,
    axis_font_size: int,
    tick_font_size: int,
    legend_font_size: int,
    value_label_font_size: int,
    panel_title_font_size: int,
    x_tick_rotation: int,
    show_grid: bool,
    legend_position: str,
) -> plt.Figure:
    plt.rcParams["font.family"] = font_family
    theme = get_theme(theme_name)

    n_panels = len(panel_variables)
    ncols = max(1, min(panel_columns, n_panels))
    nrows = math.ceil(n_panels / ncols)

    fig, axes = plt.subplots(
        nrows=nrows,
        ncols=ncols,
        figsize=(figure_width, figure_height),
        sharey=share_y_axis,
        squeeze=False,
    )
    fig.patch.set_facecolor(theme["figure_facecolor"])

    legend_handles = []
    legend_labels = []

    for panel_idx, variable in enumerate(panel_variables):
        row = panel_idx // ncols
        col = panel_idx % ncols
        ax = axes[row][col]
        ax.set_facecolor(theme["axes_facecolor"])

        pivot = pivot_for_variable(long_df, x_col=x_col, variable=variable)
        x_values = np.arange(len(pivot.index))
        x_labels = format_x_labels(pivot.index.tolist())

        for series_name in pivot.columns:
            y_values = pd.to_numeric(pivot[series_name], errors="coerce").to_numpy(dtype=float)
            color = line_colors.get(str(series_name), "#1F4E79")
            line, = ax.plot(
                x_values,
                y_values,
                label=str(series_name),
                color=color,
                linewidth=line_width,
                marker=marker_style if show_markers else None,
                markersize=marker_size if show_markers else 0,
                markeredgecolor="white",
                markeredgewidth=0.6,
            )

            if str(series_name) not in legend_labels:
                legend_handles.append(line)
                legend_labels.append(str(series_name))

            if show_value_labels:
                for x, y in zip(x_values, y_values):
                    if pd.isna(y) or not np.isfinite(y):
                        continue
                    ax.annotate(
                        format_numeric_label(y, value_label_decimal_digits, compact=compact_labels),
                        (x, y),
                        textcoords="offset points",
                        xytext=(0, 7),
                        ha="center",
                        fontsize=value_label_font_size,
                        color=theme["text_color"],
                    )

        ax.set_title(str(variable), fontsize=panel_title_font_size, fontweight="bold", color=theme["text_color"], pad=10)
        ax.set_xlabel(x_label, fontsize=axis_font_size, labelpad=8, color=theme["text_color"])
        ax.set_ylabel(y_label, fontsize=axis_font_size, labelpad=8, color=theme["text_color"])
        ax.set_xticks(x_values)
        ax.set_xticklabels(
            x_labels,
            rotation=x_tick_rotation,
            ha="right" if x_tick_rotation > 0 else "center",
            fontsize=tick_font_size,
            color=theme["text_color"],
        )
        ax.tick_params(axis="y", labelsize=tick_font_size, colors=theme["text_color"])
        apply_common_chart_style(ax, theme, show_grid)

    # Hide unused axes
    for panel_idx in range(n_panels, nrows * ncols):
        row = panel_idx // ncols
        col = panel_idx % ncols
        axes[row][col].axis("off")

    fig.suptitle(title, fontsize=title_font_size, fontweight="bold", color=theme["text_color"], y=0.995)
    if subtitle.strip():
        fig.text(0.5, 0.965, subtitle, ha="center", va="top", fontsize=subtitle_font_size, color=theme["text_color"])

    if legend_handles:
        if legend_position == "Outside right":
            legend = fig.legend(
                legend_handles,
                legend_labels,
                loc="upper left",
                bbox_to_anchor=(1.005, 0.94),
                frameon=True,
                fontsize=legend_font_size,
            )
            fig.tight_layout(rect=[0, 0, 0.86, 0.93])
        elif legend_position == "Bottom":
            legend = fig.legend(
                legend_handles,
                legend_labels,
                loc="lower center",
                bbox_to_anchor=(0.5, -0.005),
                ncol=min(4, len(legend_labels)),
                frameon=True,
                fontsize=legend_font_size,
            )
            fig.tight_layout(rect=[0, 0.06, 1, 0.93])
        else:
            legend = fig.legend(
                legend_handles,
                legend_labels,
                loc="upper right",
                frameon=True,
                fontsize=legend_font_size,
            )
            fig.tight_layout(rect=[0, 0, 1, 0.93])

        legend.get_frame().set_alpha(0.88)
        for text in legend.get_texts():
            text.set_color(theme["text_color"])
    else:
        fig.tight_layout(rect=[0, 0, 1, 0.93])

    return fig


# =====================================================
# USER INTERFACE: HEADER
# =====================================================

with st.container():
    left, right = st.columns([1, 4])
    with left:
        if LOGO_PATH.exists():
            safe_streamlit_image(LOGO_PATH, width=220)
        else:
            st.markdown("### STATCAL")
    with right:
        st.title(APP_NAME)
        st.subheader(APP_TITLE)
        st.caption(
            "A Python Streamlit application for creating publication-ready mean line charts from panel datasets. "
            "Users can upload Excel data, select numeric variables, split line charts by categorical variables, "
            "create multi-panel figures by numeric variable, and export high-resolution PNG figures suitable for "
            "academic reports, research presentations, and Scopus-indexed publications."
        )

st.markdown(
    f"""
    **Website:** [{WEBSITE_URL}]({WEBSITE_URL})  
    **Training Data / Sample Data:** [Open Google Drive Folder]({TRAINING_DATA_URL})  
    **{APP_UPDATED}**  
    **Purpose:** Upload panel data, calculate mean values for numeric variables, split trends by category, create multi-panel line charts, and export publication-ready figures.

    ---
    """
)


# =====================================================
# USER INTERFACE: DATA INPUT
# =====================================================

st.sidebar.header("1. Data Input")
st.sidebar.markdown(
    f"**Training Data / Sample Data:** [Open Google Drive Folder]({TRAINING_DATA_URL})"
)
st.sidebar.caption("Download the sample Excel dataset from Google Drive, then upload it using the menu below.")

uploaded_file = st.sidebar.file_uploader(
    "Upload Excel file",
    type=["xlsx", "xls"],
    help="Upload panel data in Excel format. The application will detect numeric variables automatically.",
)

use_sample_data = False
if uploaded_file is None and SAMPLE_DATA_PATH.exists():
    use_sample_data = st.sidebar.checkbox("Use sample data: data_sektor_energi.xlsx", value=True)

try:
    if uploaded_file is not None:
        excel_source = BytesIO(uploaded_file.getvalue())
        source_name = uploaded_file.name
    elif use_sample_data:
        excel_source = SAMPLE_DATA_PATH
        source_name = SAMPLE_DATA_PATH.name
    else:
        st.info("Please upload an Excel file to start the analysis.")
        st.stop()

    excel_file = pd.ExcelFile(excel_source)
    sheet_name = st.sidebar.selectbox("Worksheet", excel_file.sheet_names, index=0)
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    df = clean_dataframe(df)

except Exception as exc:
    st.error("Failed to read the Excel file.")
    st.exception(exc)
    st.stop()

if df.empty:
    st.error("The selected worksheet is empty.")
    st.stop()

columns = df.columns.tolist()
numeric_candidates = detect_numeric_columns(df)

if not numeric_candidates:
    st.error("No numeric variables were detected in the selected worksheet.")
    st.stop()


# =====================================================
# SIDEBAR: FILTERS
# =====================================================

st.sidebar.header("2. Flexible Data Filters")
filter_cols = st.sidebar.multiselect(
    "Select category variables for filtering",
    options=columns,
    default=default_filter_columns(columns),
    help="Examples: Company, Source Code, Year, Sector, or other categorical variables.",
)

filtered_df = apply_category_filters(df, filter_cols)
if filtered_df.empty:
    st.error("No rows remain after filtering. Please adjust the filter settings.")
    st.stop()


# =====================================================
# SIDEBAR: NUMERIC VARIABLES AND STATISTICS
# =====================================================

st.sidebar.header("3. Variables and Descriptive Statistics")

selected_numeric_cols = st.sidebar.multiselect(
    "Select numeric variables",
    options=numeric_candidates,
    default=default_numeric_columns(numeric_candidates),
    help="Select one or more numeric variables for mean line charts and descriptive statistics.",
)

if not selected_numeric_cols:
    st.warning("Please select at least one numeric variable.")
    st.stop()

decimal_digits = st.sidebar.slider("Decimal digits", min_value=0, max_value=8, value=3, step=1)

group_stats_cols = st.sidebar.multiselect(
    "Group descriptive statistics by category",
    options=columns,
    default=[],
    help="Examples: Year, Company, Source Code, Sector. Leave blank for overall descriptive statistics.",
)

descriptive_stats = compute_descriptive_statistics(
    df=filtered_df,
    numeric_cols=selected_numeric_cols,
    decimal_digits=decimal_digits,
)

grouped_descriptive_stats = compute_grouped_descriptive_statistics(
    df=filtered_df,
    numeric_cols=selected_numeric_cols,
    group_cols=group_stats_cols,
    decimal_digits=decimal_digits,
)


# =====================================================
# SIDEBAR: MEAN LINE CHART SETTINGS
# =====================================================

st.sidebar.header("4. Mean Line Chart Settings")

x_index = preferred_option(columns, ["Year", "Tahun", "Date", "Time", "Period", "Periode"], fallback_index=0)
x_col = st.sidebar.selectbox("X-axis variable", columns, index=x_index)

split_options = ["None"] + columns
split_index = preferred_option(split_options, ["Company", "Source Code", "Year", "Sector", "Sektor"], fallback_index=0)
split_col = st.sidebar.selectbox(
    "Split lines by category",
    split_options,
    index=split_index,
    help="Examples: Company or Year. Each category will become a separate line inside every numeric-variable panel.",
)

sort_x = st.sidebar.checkbox("Sort X-axis", value=True)

max_panels = st.sidebar.slider("Maximum numeric-variable panels", 1, 20, min(6, len(selected_numeric_cols)))
panel_variables = selected_numeric_cols[:max_panels]
panel_columns = st.sidebar.slider("Number of panel columns", 1, 4, min(2, max_panels))
share_y_axis = st.sidebar.checkbox("Share Y-axis across panels", value=False)

chart_long_data = build_mean_line_long_data(
    df=filtered_df,
    x_col=x_col,
    numeric_cols=panel_variables,
    split_col=split_col,
    sort_x=sort_x,
)

if chart_long_data.empty:
    st.error("The mean line chart data is empty. Please adjust the X-axis, numeric variables, split variable, or filters.")
    st.stop()


# =====================================================
# SIDEBAR: PUBLICATION STYLE SETTINGS
# =====================================================

st.sidebar.header("5. Publication Style Settings")

font_family = st.sidebar.selectbox("Font family", ["Arial", "Times New Roman", "DejaVu Sans", "DejaVu Serif"], index=2)
theme_name = st.sidebar.selectbox("Chart background theme", list(THEMES.keys()), index=0)
palette_name = st.sidebar.selectbox("Color palette", list(COLOR_PALETTES.keys()), index=0)

chart_title = st.sidebar.text_area("Chart title", value="Mean Trend of Selected Numeric Variables", height=68)
chart_subtitle = st.sidebar.text_input("Chart subtitle", value="Panel chart based on mean values from filtered panel data")
x_axis_label = st.sidebar.text_input("X-axis label", value=x_col)
y_axis_label = st.sidebar.text_input("Y-axis label", value="Mean")

figure_width = st.sidebar.slider("Figure width", 6.0, 30.0, 14.0, 0.5)
figure_height = st.sidebar.slider("Figure height", 4.0, 30.0, 10.0, 0.5)

line_width = st.sidebar.slider("Line width", 0.5, 8.0, 2.4, 0.1)
show_markers = st.sidebar.checkbox("Add dot markers to lines", value=True)
marker_style = st.sidebar.selectbox("Marker style", MARKERS, index=0)
marker_size = st.sidebar.slider("Marker size", 2, 18, 6)
show_value_labels = st.sidebar.checkbox("Show mean values on the chart", value=False)
compact_value_labels = st.sidebar.checkbox("Use compact value labels (K/M/B/T)", value=True)
x_tick_rotation = st.sidebar.slider("X-axis label rotation", 0, 90, 0, 5)
show_grid = st.sidebar.checkbox("Show grid", value=True)
legend_position = st.sidebar.selectbox("Legend position", ["Outside right", "Bottom", "Best"], index=0)

common_title_font_size = st.sidebar.slider("Main title font size", 10, 44, 18)
common_subtitle_font_size = st.sidebar.slider("Subtitle font size", 8, 30, 11)
panel_title_font_size = st.sidebar.slider("Panel title font size", 8, 30, 13)
common_axis_font_size = st.sidebar.slider("Axis label font size", 8, 30, 12)
common_tick_font_size = st.sidebar.slider("Tick label font size", 6, 26, 10)
common_legend_font_size = st.sidebar.slider("Legend font size", 6, 24, 9)
common_value_label_font_size = st.sidebar.slider("Value label font size", 5, 22, 8)

# Color settings are based on line categories. If there is no split variable,
# each numeric variable appears as one line in its own panel.
line_series = sorted(chart_long_data["Split"].dropna().astype(str).unique().tolist(), key=lambda x: str(x).lower())
palette = get_palette_color_list(palette_name)
line_colors: Dict[str, str] = {}
st.sidebar.markdown("**Line colors**")
for idx, series_name in enumerate(line_series[:30]):
    line_colors[str(series_name)] = st.sidebar.color_picker(
        f"Line color: {series_name}",
        palette[idx % len(palette)],
    )
if len(line_series) > 30:
    st.sidebar.warning("Only the first 30 line colors are shown for customization.")


# =====================================================
# MAIN DISPLAY: TAB LAYOUT AND EXPORT SETTINGS
# =====================================================

# The application uses tabs in the main page to make the workflow easier to follow.
# Export settings are intentionally placed in the last tab.

mean_line_fig = create_panel_mean_line_chart(
    long_df=chart_long_data,
    x_col=x_col,
    panel_variables=panel_variables,
    title=chart_title,
    subtitle=chart_subtitle,
    x_label=x_axis_label,
    y_label=y_axis_label,
    figure_width=figure_width,
    figure_height=figure_height,
    panel_columns=panel_columns,
    share_y_axis=share_y_axis,
    theme_name=theme_name,
    font_family=font_family,
    line_colors=line_colors,
    line_width=line_width,
    show_markers=show_markers,
    marker_style=marker_style,
    marker_size=marker_size,
    show_value_labels=show_value_labels,
    value_label_decimal_digits=decimal_digits,
    compact_labels=compact_value_labels,
    title_font_size=common_title_font_size,
    subtitle_font_size=common_subtitle_font_size,
    axis_font_size=common_axis_font_size,
    tick_font_size=common_tick_font_size,
    legend_font_size=common_legend_font_size,
    value_label_font_size=common_value_label_font_size,
    panel_title_font_size=panel_title_font_size,
    x_tick_rotation=x_tick_rotation,
    show_grid=show_grid,
    legend_position=legend_position,
)

tab_data, tab_stats, tab_grouped_stats, tab_chart_data, tab_chart, tab_export = st.tabs(
    [
        "1. Dataset Preview",
        "2. Descriptive Statistics",
        "3. Grouped Statistics",
        "4. Chart Data",
        "5. Mean Line Panel Chart",
        "6. Export Settings",
    ]
)


with tab_data:
    st.subheader("Dataset Preview")
    st.write(f"**Source:** {source_name} | **Worksheet:** {sheet_name}")
    st.dataframe(make_arrow_safe_dataframe(df.head(50)))

    metric_1, metric_2, metric_3, metric_4 = st.columns(4)
    with metric_1:
        st.metric("Rows after filtering", f"{len(filtered_df):,}")
    with metric_2:
        st.metric("Columns", f"{len(df.columns):,}")
    with metric_3:
        st.metric("Selected numeric variables", f"{len(selected_numeric_cols):,}")
    with metric_4:
        st.metric("Line split variable", split_col)

    with st.expander("Filtered data preview", expanded=False):
        st.dataframe(make_arrow_safe_dataframe(filtered_df.head(100)))

    with st.expander("Detected numeric variables", expanded=False):
        st.write(numeric_candidates)


with tab_stats:
    st.subheader("Descriptive Statistics")
    st.write("The table below displays minimum, maximum, mean, and standard deviation for the selected numeric variables.")
    st.dataframe(make_arrow_safe_dataframe(descriptive_stats))


with tab_grouped_stats:
    st.subheader("Grouped Descriptive Statistics")
    if group_stats_cols:
        st.write(f"Grouped by: {', '.join(group_stats_cols)}")
        st.dataframe(make_arrow_safe_dataframe(grouped_descriptive_stats.head(500)))
    else:
        st.info("No grouping variable was selected. Use the sidebar option 'Group descriptive statistics by category' to create grouped descriptive statistics.")
        st.dataframe(make_arrow_safe_dataframe(grouped_descriptive_stats))


with tab_chart_data:
    st.subheader("Mean Line Chart Data")
    st.write("This table contains the mean values used to build the multi-panel line chart.")
    st.dataframe(make_arrow_safe_dataframe(chart_long_data.head(500)))


with tab_chart:
    st.subheader("Publication-Ready Mean Line Panel Chart")

    if len(panel_variables) < len(selected_numeric_cols):
        st.info(
            f"Showing the first {len(panel_variables)} numeric-variable panels. "
            "Increase 'Maximum numeric-variable panels' in the sidebar to show more."
        )

    st.pyplot(mean_line_fig)


with tab_export:
    st.subheader("Export Settings")
    st.write("Use this final tab to control PNG export settings and download publication-ready outputs.")

    dpi = st.selectbox(
        "PNG resolution / DPI",
        [300, 600, 900, 1200, 1500],
        index=3,
        help="Default is 1200 DPI for high-resolution publication output.",
    )
    transparent_background = st.checkbox("Transparent PNG background", value=False)

    st.info(
        "The default PNG resolution is set to 1200 DPI. "
        "Higher DPI values may require more processing time and memory, especially for large multi-panel figures."
    )

    mean_line_png = fig_to_png_bytes(
        mean_line_fig,
        dpi=dpi,
        transparent_background=transparent_background,
    )
    st.download_button(
        label=f"⬇️ Download Mean Line Panel Chart PNG ({dpi} DPI)",
        data=mean_line_png,
        file_name="statcal_online_mean_line_panel_chart.png",
        mime="image/png",
    )

    st.markdown("---")
    st.subheader("Export Tables")
    metadata = {
        "Application": f"{APP_NAME} - {APP_TITLE}",
        "Website": WEBSITE_URL,
        "Training Data URL": TRAINING_DATA_URL,
        "Source File": source_name,
        "Worksheet": sheet_name,
        "Rows in Original Data": len(df),
        "Rows after Filtering": len(filtered_df),
        "X-axis Variable": x_col,
        "Split Lines By": split_col,
        "Selected Numeric Variables": ", ".join(selected_numeric_cols),
        "Panel Variables Shown": ", ".join(panel_variables),
        "Decimal Digits": decimal_digits,
        "PNG DPI": dpi,
        "Transparent PNG Background": transparent_background,
        "Updated": APP_UPDATED,
    }

    excel_bytes = dataframe_to_excel_bytes(
        descriptive_stats=descriptive_stats,
        grouped_descriptive_stats=grouped_descriptive_stats,
        chart_long_data=chart_long_data,
        filtered_df=filtered_df,
        metadata=metadata,
    )

    st.download_button(
        label="⬇️ Download Descriptive Statistics and Mean Line Chart Data to Excel",
        data=excel_bytes,
        file_name="statcal_online_mean_line_panel_chart_tables.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

plt.close(mean_line_fig)
